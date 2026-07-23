"""
Import/export helpers for the manager's task spreadsheet.

Column matching is intentionally forgiving: real spreadsheets people type by
hand rarely match a template exactly (typos, merged wording, stray spaces,
parentheses jammed against words). We normalize headers and match them with
ordered "contains" rules instead of requiring an exact string.

Date parsing is equally forgiving: it accepts native Excel date cells,
Excel's numeric date serials, ISO dates, and common hand-typed formats
including the everyday Indian day-first format (DD-MM-YYYY / DD/MM/YYYY).
"""
import csv
import io
import re
from datetime import date, datetime

import openpyxl

MONTH_NAMES = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "sept": 9, "oct": 10, "nov": 11, "dec": 12,
}

# Ordered rules: first match wins. More specific fields are checked before
# the generic "date" catch-all so a header like "Date Type (Static/Reel/Carousel)"
# still resolves to Content Type rather than Date.
HEADER_RULES = [
    ("deadline", lambda h: "deadline" in h),
    ("postingType", lambda h: "postingtype" in h or "posting" in h or "storyfeed" in h),
    ("contentType", lambda h: "contenttype" in h or h.startswith("type") or "staticreel" in h or "reelcarousel" in h),
    ("objective", lambda h: "objective" in h),
    ("details", lambda h: "detail" in h),
    ("caption", lambda h: "caption" in h),
    ("reference", lambda h: "reference" in h or "ref" in h),
    ("remark", lambda h: "remark" in h),
    ("date", lambda h: "date" in h),
]
FIELD_LABELS = {
    "date": "Date", "deadline": "Deadline", "contentType": "Type",
    "postingType": "Posting Type", "objective": "Objective", "details": "Details",
    "caption": "Caption", "reference": "Reference", "remark": "Remark",
}

TEMPLATE_HEADERS = [
    "Date", "Type (Static/Reel/Carousel)", "Posting Type (Story/Feed)",
    "Objective", "Details", "Caption", "Reference", "Deadline", "Remark",
]
TEMPLATE_SAMPLE = [
    "22-07-2026", "Reel", "Feed", "Awareness",
    "Short reel on new product line", "See you this weekend!",
    "brand-guidelines.pdf", "25-07-2026", "Priority client",
]


def normalize_header(h):
    return re.sub(r"[^a-z0-9]", "", str(h or "").lower())


def match_columns(header_row):
    """Return {field: column_index} using the ordered header rules."""
    col_index = {}
    for i, raw in enumerate(header_row):
        h = normalize_header(raw)
        if not h:
            continue
        for field, test in HEADER_RULES:
            if field not in col_index and test(h):
                col_index[field] = i
                break
    return col_index


def parse_flexible_date(value):
    """Return an ISO yyyy-mm-dd string, or None if it can't be parsed."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    s = str(value).strip()
    if not s:
        return None

    m = re.match(r"^(\d{4})[/\-](\d{1,2})[/\-](\d{1,2})$", s)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return _safe_date(y, mo, d)

    m = re.match(r"^(\d{1,2})[\s\-/]([a-zA-Z]{3,9})[\s\-/,]+(\d{2,4})$", s)
    if m:
        mo = MONTH_NAMES.get(m.group(2)[:4].lower().replace("sept", "sep")) or MONTH_NAMES.get(m.group(2)[:3].lower())
        if mo:
            y = int(m.group(3))
            y = y + 2000 if y < 100 else y
            return _safe_date(y, mo, int(m.group(1)))

    m = re.match(r"^([a-zA-Z]{3,9})[\s\-/,]+(\d{1,2})[\s\-/,]+(\d{2,4})$", s)
    if m:
        mo = MONTH_NAMES.get(m.group(1)[:4].lower().replace("sept", "sep")) or MONTH_NAMES.get(m.group(1)[:3].lower())
        if mo:
            y = int(m.group(3))
            y = y + 2000 if y < 100 else y
            return _safe_date(y, mo, int(m.group(2)))

    m = re.match(r"^(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{2,4})$", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        y = y + 2000 if y < 100 else y
        if mo > 12 and d <= 12:
            d, mo = mo, d
        return _safe_date(y, mo, d)

    return None


def _safe_date(y, mo, d):
    try:
        return date(y, mo, d).isoformat()
    except ValueError:
        return None


def read_rows_from_upload(file_storage):
    """Return a list-of-lists (first row = headers) from an uploaded .xlsx or .csv file."""
    filename = (file_storage.filename or "").lower()
    raw = file_storage.read()

    if filename.endswith(".csv"):
        text = raw.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        return [row for row in reader]

    # default: treat as xlsx (also covers files with no/odd extension)
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    ws = wb.worksheets[0]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    return rows


def build_template_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tasks"
    ws.append(TEMPLATE_HEADERS)
    ws.append(TEMPLATE_SAMPLE)
    for col_cells in ws.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 12), 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def import_tasks(rows):
    """
    Parse spreadsheet rows into task dicts.
    Returns (tasks, skipped_count, skipped_row_numbers, error, detected_headers)
    """
    if not rows:
        return [], 0, [], "That file has no rows.", []

    header_row = rows[0]
    col_index = match_columns(header_row)
    detected_headers = [str(h).strip() for h in header_row if str(h or "").strip()]

    missing = [f for f in ("date", "deadline") if f not in col_index]
    if missing:
        label = " and ".join(FIELD_LABELS[f] for f in missing)
        return [], 0, [], f"Couldn't find a column for {label}.", detected_headers

    def get(row, field):
        idx = col_index.get(field)
        if idx is None or idx >= len(row):
            return ""
        val = row[idx]
        return val if val is not None else ""

    tasks = []
    skipped = 0
    skipped_rows = []

    for r in range(1, len(rows)):
        row = rows[r]
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue

        task_date = parse_flexible_date(get(row, "date"))
        deadline = parse_flexible_date(get(row, "deadline"))
        if not task_date or not deadline:
            skipped += 1
            skipped_rows.append(r + 1)
            continue

        tasks.append({
            "date": task_date,
            "deadline": deadline,
            "contentType": str(get(row, "contentType") or "").strip() or "Static",
            "postingType": str(get(row, "postingType") or "").strip() or "Feed",
            "objective": str(get(row, "objective") or "").strip(),
            "details": str(get(row, "details") or "").strip(),
            "caption": str(get(row, "caption") or "").strip(),
            "reference": str(get(row, "reference") or "").strip(),
            "remark": str(get(row, "remark") or "").strip(),
        })

    return tasks, skipped, skipped_rows, None, detected_headers
